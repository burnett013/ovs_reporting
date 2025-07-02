# parser/ug_parser.py
from pathlib import Path
from PyPDF2 import PdfReader
import pandas as pd
import re
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def is_accredited(lines: list) -> str:
    text = " ".join(lines).lower()
    keywords = [
        "not accredited", "no accreditation", "accreditation is not required",
        "not eligible for accreditation", "does not hold accreditation", "unaccredited"
    ]
    return "No" if any(k in text for k in keywords) else "Yes"

def get_program_pid(formatted_title: str, pid_df: pd.DataFrame) -> str:
    match = pid_df[pid_df["Program"].str.lower().str.strip() == formatted_title.lower().strip()]
    return match.iloc[0]["PID"] if not match.empty else ""

def extract_catalog_page_number(lines: list) -> int:
    for line in reversed(lines):
        if line.strip().isdigit() and len(line.strip()) >= 3:
            return int(line.strip())
    return None

def format_program_name(name: str) -> str:
    title_cased = name.title()
    corrections = {
        "B.A.": "B.A.", "B.S.": "B.S.", "Ph.D.": "Ph.D.",
        "With": "with", "Rotc": "ROTC", "Esol": "ESOL",
        "And": "and", "'S": "'s", "Gpa": "GPA"
    }
    for wrong, right in corrections.items():
        title_cased = re.sub(rf"\b{wrong}\b", right, title_cased)
    return title_cased

def extract_modality_from_lines(lines: list) -> str:
    text = " ".join(lines).lower()
    if re.search(r"(fully|100%)\s+online", text) or any(k in text for k in ["offered online", "delivered online", "available online", "online format"]):
        return "Online"
    if any(k in text for k in ["hybrid", "blended", "online and on campus"]):
        return "Hybrid"
    if any(k in text for k in ["on campus only", "in-person only", "campus-based"]):
        return "Campus"
    if "online" in text:
        return "Online"
    if "on campus" in text:
        return "Campus"
    return "Campus"

def has_license_prep(lines: list) -> str:
    text = " ".join(lines).lower()
    keywords = [
        "state-approved program", "leads to certification", "eligible for teacher certification",
        "teacher preparation program", "florida teacher certification exam", "eligible for the endorsements"
    ]
    return "Yes" if any(k in text for k in keywords) else "No"

def extract_credit_hours_from_line(line: str) -> int:
    patterns = [
        r"TOTAL\s+(DEGREE|MAJOR|CERTIFICATE)\s+HOURS\s*:\s*(\d+)",
        r"TOTAL\s+HOURS\s*:\s*(\d+)"
    ]
    for pattern in patterns:
        match = re.search(pattern, line.upper())
        if match:
            return int(match.group(match.lastindex))
    return None

def extract_major_credit_hours(lines: list) -> int:
    return next((extract_credit_hours_from_line(line) for line in lines if extract_credit_hours_from_line(line)), None)

extract_concentration_credit_hours = extract_major_credit_hours

def extract_certificate_credit_hours(lines: list) -> int:
    patterns = [
        r"TOTAL\s+CERTIFICATE\s+HOURS\s*[:\-]?\s*(\d+)",
        r"CERTIFICATE\s+CORE\s*\((\d+)\s+CREDIT\s+HOURS\)",
        r"CERTIFICATE\s+CORE\s+COURSES\s*\((\d+)\s+CREDIT\s+HOURS\)",
        r"CERTIFICATE\s+REQUIREMENTS\s*[:\-]?\s*(\d+)\s+CREDIT\s+HOURS",
        r"(\d+)\s+CREDIT\s+HOURS\s+REQUIRED"
    ]
    for line in lines:
        for pattern in patterns:
            match = re.search(pattern, line.upper())
            if match:
                return int(match.group(1))
    return None

def extract_minor_credit_hours(lines: list) -> int:
    patterns = [
        r"TOTAL\s+MINOR\s+(?:CREDIT\s+)?HOURS\s*[:\-]?\s*(\d+)",
        r"REQUIRES\s+A\s+TOTAL\s+OF\s+(\d+)\s+CREDIT\s+HOURS",
        r"COMPLETION\s+OF\s+THE\s+MINOR\s+REQUIRES\s+(\d+)\s+CREDIT\s+HOURS",
        r"CONSISTS\s+OF\s+A\s+MINIMUM\s+OF\s+(\d+)\s+CREDIT\s+HOURS",
        r"MINOR\s+(?:CORE|REQUIRED|ELECTIVE)?\s*(?:COURSES)?\s*\((\d+)\s+CREDIT\s+HOURS\)"
    ]
    values = {int(m.group(1)) for line in lines for p in patterns if (m := re.search(p, line, flags=re.IGNORECASE))}
    if values: return max(values)

    total = 0
    seen = set()
    component_patterns = [
        r"MINOR\s+CORE\s+CREDIT\s+HOURS\s*[:\-]?\s*(\d+)",
        r"MINOR\s+ELECTIVE\s+CREDIT\s+HOURS\s*[:\-]?\s*(\d+)"
    ]
    for line in lines:
        for pattern in component_patterns:
            match = re.search(pattern, line, flags=re.IGNORECASE)
            if match:
                value = int(match.group(1))
                if (line, value) not in seen:
                    seen.add((line, value))
                    total += value
    return total or None

def is_valid_program_name(line: str) -> bool:
    suffixes = ["B.S.", "B.A.", "MINOR", "CERTIFICATE", "CONCENTRATION"]
    invalid_prefixes = [
        "GRADUATION REQUIREMENTS", "RESEARCH OPPORTUNITIES", "ADVISING INFORMATION",
        "STATE MANDATED", "STATE MATHEMATICS", "MAJOR CORE", "REQUIRED COURSES",
        "REQUIREMENTS", "OTHER REQUIREMENTS"
    ]
    upper = line.upper()
    return not any(upper.startswith(p) for p in invalid_prefixes) and any(s in upper for s in suffixes)

def classify_program_type(name: str) -> str:
    upper = name.upper()
    if "MINOR" in upper: return "Minor"
    if "CERTIFICATE" in upper: return "Certificate"
    if "CONCENTRATION" in upper: return "Concentration"
    if "B.A." in upper or "B.S." in upper: return "Major"
    return "Unknown"

def extract_program_names(pdf_path: Path) -> list:
    reader = PdfReader(pdf_path)
    results = []

    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if not text:
            continue

        lines = [line.strip() for line in text.split('\n') if line.strip()]
        page_num = extract_catalog_page_number(lines)
        if not page_num or page_num <= 145:
            continue

        for j, line in enumerate(lines):
            if "UNIVERSITY OF SOUTH FLORIDA" in line.upper() and "UNDERGRADUATE CATALOG" in line.upper():
                title_lines = []
                for k in range(j + 1, len(lines)):
                    next_line = lines[k]
                    if not next_line or re.search(r"[a-z]", next_line): break
                    if "TOTAL DEGREE HOURS" in next_line.upper(): break
                    title_lines.append(next_line)

                full_title = " ".join(title_lines)
                if full_title and is_valid_program_name(full_title):
                    name = format_program_name(full_title)
                    name = re.sub(r"^\d+\s+", "", name)
                    name = re.sub(r"\s*Total\s+(Minor|Major|Certificate)?\s*Hours:\s*\d+", "", name, flags=re.IGNORECASE)
                    name = re.sub(r"\s*Minor Requirements$", "", name, flags=re.IGNORECASE)
                    name = re.sub(r"\(\s*\d+\s+Credit\s+Hours\s*\)", "", name, flags=re.IGNORECASE)
                    name = re.sub(r"\s*-\s*\d+\s*$", "", name).strip()

                    if any(x in name.upper() for x in [
                        "STATE MANDATED", "ADDITIONAL INFORMATION", "PROGRESSION REQUIREMENTS",
                        "ADVISING INFORMATION", "STATE MATHEMATICS PATHWAY", "RESEARCH OPPORTUNITIES",
                        "TRAINING OPTION HTTPS", "TWO SPC", "CREDIT HOURS CONCENTRATION",
                        "CONCENTRATION CORE", "ELECTIVE COURSES", "MINOR MINOR", "CONCENTRATION CORE COURSE",
                        "INTERNSHIP OPPORTUNITIES", "RESPONSIBLE AND INCLUSIVE", "CONCENTRATION REQUIREMENT"
                    ]):
                        continue

                    program_type = classify_program_type(name)
                    block = lines[j+1 : j+75]
                    credit = {
                        "Major": extract_major_credit_hours,
                        "Concentration": extract_concentration_credit_hours,
                        "Certificate": extract_certificate_credit_hours,
                        "Minor": extract_minor_credit_hours
                    }.get(program_type, lambda _: None)(block)

                    edu = {"Major": "Bachelor's", "Minor": "Bachelor's", "Concentration": "Bachelor's", "Certificate": "Certificate"}.get(program_type, "Unknown")
                    modality = extract_modality_from_lines(lines[j+1 : j+20])
                    license_prep = has_license_prep(block)
                    accredited = is_accredited(block)

                    results.append({
                        "PID": "",
                        "Program Name": name,
                        "Accredited": accredited,
                        "Type": program_type,
                        "Concentrations? Yes or No": "Yes" if program_type == "Concentration" else "No",
                        "Total Credit Hours in Program": credit,
                        "Program Length Measurement": "Semester",
                        "Full-Time Enrollment": 12,
                        "Page Number": page_num,
                        "Educational Objective": edu,
                        "License Prep": license_prep,
                        "Modality": modality,
                    })
                break
    return results

def export_to_excel(data: list, output_path: Path) -> pd.DataFrame:
    if not data:
        print("No valid program data found.")
        return pd.DataFrame()

    df = pd.DataFrame(data)
    df.insert(0, "Program Name", df.pop("Program Name"))
    df = df[[
        "Program Name", "Accredited", "Educational Objective", "Concentrations? Yes or No",
        "Total Credit Hours in Program", "Program Length Measurement", "Full-Time Enrollment",
        "Page Number", "License Prep", "Modality", "Type"
    ]]

    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    widths = {
        1: 15, 2: 50, 3: 15, 4: 25, 5: 20, 6: 25, 7: 15, 8: 25,
        9: 20, 10: 20, 11: 25
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)
    print(f"\nExtracted {len(data)} programs")
    return df

def run_ug_parser(input_pdf: str) -> pd.DataFrame:
    program_data = extract_program_names(Path(input_pdf))
    df = pd.DataFrame(program_data)
    return df